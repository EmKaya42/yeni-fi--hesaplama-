from __future__ import annotations

import os
import re
import sqlite3
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any

from flask import Flask, jsonify, render_template, request, send_file
from werkzeug.utils import secure_filename
from services.extraction_service import MISSING, extract_receipt, extract_z_report
from services.ocr_service import read_image

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "data" / "fis_takip.db"
UPLOAD_DIR = BASE_DIR / "data" / "uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
OCR_READER = None

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 12 * 1024 * 1024

RECEIPT_FIELDS = [
    "product_name", "tax_id", "invoice_no", "receipt_no", "receipt_datetime",
    "vat_rate", "vat_base", "vat_amount", "total_amount",
]
Z_FIELDS = [
    "report_no", "report_datetime", "daily_turnover", "vat1_base", "vat1_amount",
    "vat10_base", "vat10_amount", "vat20_base", "vat20_amount", "cash_amount",
    "card_amount", "other_payment", "cancel_amount", "refund_amount",
    "discount_amount", "first_receipt_no", "last_receipt_no",
]
Z_EXTRA_FIELDS = ["product_name", "tax_id", "invoice_no"]
Z_INPUT_FIELDS = Z_FIELDS + Z_EXTRA_FIELDS


def get_db() -> sqlite3.Connection:
    connection = sqlite3.connect(DB_PATH)
    connection.row_factory = sqlite3.Row
    return connection


def init_db() -> None:
    with get_db() as connection:
        connection.execute("""
            CREATE TABLE IF NOT EXISTS receipts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_name TEXT NOT NULL, tax_id TEXT, invoice_no TEXT,
                receipt_no TEXT, receipt_datetime TEXT, vat_rate REAL,
                vat_base REAL, vat_amount REAL, total_amount REAL NOT NULL,
                created_at TEXT NOT NULL, user_id TEXT NOT NULL DEFAULT 'legacy'
            )
        """)
        connection.execute("""
            CREATE TABLE IF NOT EXISTS z_reports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                report_no TEXT NOT NULL, report_datetime TEXT,
                daily_turnover REAL NOT NULL, vat1_base REAL, vat1_amount REAL,
                vat10_base REAL, vat10_amount REAL, vat20_base REAL, vat20_amount REAL,
                cash_amount REAL, card_amount REAL, other_payment REAL,
                cancel_amount REAL, refund_amount REAL, discount_amount REAL,
                first_receipt_no TEXT, last_receipt_no TEXT, product_name TEXT,
                tax_id TEXT, invoice_no TEXT, created_at TEXT NOT NULL, user_id TEXT NOT NULL DEFAULT 'legacy'
            )
        """)
        columns = {row[1] for row in connection.execute("PRAGMA table_info(z_reports)")}
        receipt_columns = {row[1] for row in connection.execute("PRAGMA table_info(receipts)")}
        for column in ("user_id",):
            if column not in receipt_columns:
                connection.execute("ALTER TABLE receipts ADD COLUMN user_id TEXT NOT NULL DEFAULT 'legacy'")
        for column in ("product_name", "tax_id", "invoice_no", "user_id"):
            if column not in columns:
                connection.execute(f"ALTER TABLE z_reports ADD COLUMN {column} TEXT")


def current_user_id() -> str:
    return request.headers.get("X-Firebase-UID", "local-user").strip() or "local-user"


def number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    cleaned = str(value).strip().replace("₺", "").replace("TL", "")
    cleaned = cleaned.replace(".", "").replace(",", ".")
    try:
        return round(float(re.sub(r"[^0-9.-]", "", cleaned)), 2)
    except ValueError:
        return 0.0


def payload(fields: list[str]) -> dict[str, Any]:
    data = request.get_json(silent=True) or request.form.to_dict()
    return {field: data.get(field, "") for field in fields}


def row_dict(row: sqlite3.Row) -> dict[str, Any]:
    return dict(row)


def ocr_text(image_field: str) -> tuple[str, str | None]:
    image = request.files.get(image_field)
    if not image or not image.filename:
        return "", "Görsel seçilmedi."
    try:
        image_path = Path(tempfile.gettempdir()) / f"fis_ocr_{datetime.now().strftime('%Y%m%d%H%M%S%f')}_{secure_filename(image.filename)}"
        image.save(image_path)
        text, engine = read_image(image_path)
        image_path.unlink(missing_ok=True)
        return text, None if text else f"{engine} görselde okunabilir metin bulamadı."
    except RuntimeError as error:
        return "", str(error)
    except Exception as error:
        return "", f"OCR okunamadı: {error}"


def find_value(text: str, labels: str) -> str:
    match = re.search(rf"(?:{labels})\s*[:#-]?\s*([^\n]+)", text, re.IGNORECASE)
    return match.group(1).strip() if match else ""


def parse_receipt(text: str) -> dict[str, Any]:
    money = r"[0-9][0-9.,]*"
    vat_rate = re.search(r"(?:KDV|VAT)\s*%?\s*(1|10|20)", text, re.IGNORECASE)
    total = re.search(rf"(?:TOPLAM|GENEL TOPLAM|TOTAL)\s*[:=]?\s*({money})", text, re.IGNORECASE)
    vat_amount = re.search(rf"(?:KDV TUTARI|VAT AMOUNT)\s*[:=]?\s*({money})", text, re.IGNORECASE)
    vat_base = re.search(rf"(?:KDV MATRAHI|MATRAH)\s*[:=]?\s*({money})", text, re.IGNORECASE)
    tax_id = re.search(r"\b(?:VKN|TCKN)\s*[:#-]?\s*([0-9]{10,11})\b", text, re.IGNORECASE)
    return {
        "product_name": find_value(text, r"(?:Ürün|Hizmet|Product|Item)") or "",
        "tax_id": tax_id.group(1) if tax_id else "",
        "invoice_no": find_value(text, r"(?:Fatura|Belge|Invoice)\s*(?:No|Numarası)?") ,
        "receipt_no": find_value(text, r"(?:Fiş|Receipt)\s*(?:No|Numarası)?"),
        "receipt_datetime": find_value(text, r"(?:Tarih|Date|Saat|Time)") ,
        "vat_rate": vat_rate.group(1) if vat_rate else "",
        "vat_base": vat_base.group(1) if vat_base else "",
        "vat_amount": vat_amount.group(1) if vat_amount else "",
        "total_amount": total.group(1) if total else "",
        "raw_text": text,
    }


def parse_z_report(text: str) -> dict[str, Any]:
    result: dict[str, Any] = {field: "" for field in Z_FIELDS}
    aliases = {
        "report_no": r"(?:Z Raporu|Z Report)\s*(?:No|Numarası)?",
        "report_datetime": r"(?:Tarih|Date|Saat|Time)",
        "daily_turnover": r"(?:Günlük Toplam Ciro|Toplam Ciro|Daily Turnover)",
        "cash_amount": r"(?:Nakit|Cash)", "card_amount": r"(?:Kredi Kartı|POS|Card)",
        "other_payment": r"(?:Diğer Ödeme|Other Payment)", "cancel_amount": r"(?:İptal|Cancel)",
        "refund_amount": r"(?:İade|Refund)", "discount_amount": r"(?:İskonto|Discount)",
        "first_receipt_no": r"(?:İlk Fiş|First Receipt)", "last_receipt_no": r"(?:Son Fiş|Last Receipt)",
    }
    for field, label in aliases.items():
        result[field] = find_value(text, label)
    for rate in (1, 10, 20):
        result[f"vat{rate}_base"] = find_value(text, rf"(?:%\s*{rate}|KDV\s*{rate})\s*(?:Matrah|Base)")
        result[f"vat{rate}_amount"] = find_value(text, rf"(?:%\s*{rate}|KDV\s*{rate})\s*(?:Tutar|Amount)")
    result["raw_text"] = text
    return result


@app.get("/")
def index():
    return send_file(BASE_DIR / "index.html")


@app.get("/firebase-config.js")
def firebase_config():
    return send_file(BASE_DIR / "firebase-config.js", mimetype="application/javascript")


@app.get("/app")
def dashboard_app():
    return render_template("index.html")


@app.get("/api/dashboard")
def dashboard():
    user_id = current_user_id()
    with get_db() as connection:
        receipt = connection.execute("SELECT COALESCE(SUM(total_amount),0) AS turnover, COALESCE(SUM(vat_amount),0) AS vat, COUNT(*) AS count FROM receipts WHERE user_id=?", (user_id,)).fetchone()
        z = connection.execute("SELECT COALESCE(SUM(daily_turnover),0) AS turnover, COALESCE(SUM(vat1_amount + vat10_amount + vat20_amount),0) AS vat, COALESCE(SUM(cash_amount),0) AS cash, COALESCE(SUM(card_amount),0) AS card, COALESCE(SUM(refund_amount),0) AS refund, COALESCE(SUM(discount_amount),0) AS discount, COUNT(*) AS count FROM z_reports WHERE user_id=?", (user_id,)).fetchone()
        today = datetime.now().strftime("%Y-%m-%d")
        month = datetime.now().strftime("%Y-%m")
        receipt_periods = connection.execute("SELECT COALESCE(SUM(CASE WHEN substr(created_at,1,10)=? THEN total_amount ELSE 0 END),0) AS daily, COALESCE(SUM(CASE WHEN substr(created_at,1,7)=? THEN total_amount ELSE 0 END),0) AS monthly FROM receipts WHERE user_id=?", (today, month, user_id)).fetchone()
        z_periods = connection.execute("SELECT COALESCE(SUM(CASE WHEN substr(created_at,1,10)=? THEN daily_turnover ELSE 0 END),0) AS daily, COALESCE(SUM(CASE WHEN substr(created_at,1,7)=? THEN daily_turnover ELSE 0 END),0) AS monthly FROM z_reports WHERE user_id=?", (today, month, user_id)).fetchone()
    return jsonify({"receipts": row_dict(receipt), "z_reports": row_dict(z), "periods": {"daily": receipt_periods["daily"] + z_periods["daily"], "monthly": receipt_periods["monthly"] + z_periods["monthly"]}})


@app.get("/api/receipts")
def receipts():
    user_id = current_user_id()
    with get_db() as connection:
        rows = connection.execute("SELECT * FROM receipts WHERE user_id=? ORDER BY id DESC", (user_id,)).fetchall()
    return jsonify([row_dict(row) for row in rows])


@app.post("/api/receipts")
def create_receipt():
    user_id = current_user_id()
    data = payload(RECEIPT_FIELDS)
    if not data["product_name"] or data["product_name"] == MISSING:
        data["product_name"] = "Bulunamadı"
    if data["total_amount"] in {"", MISSING}:
        return jsonify({"error": "Ürün/hizmet adı ve toplam tutar zorunludur."}), 400
    with get_db() as connection:
        duplicate = connection.execute("SELECT id FROM receipts WHERE user_id=? AND receipt_no = ? AND receipt_no NOT IN ('', ?) AND total_amount = ?", (user_id, data["receipt_no"], MISSING, number(data["total_amount"]))).fetchone()
        if duplicate:
            return jsonify({"ok": True, "duplicate": True, "message": "Bu fiş daha önce kaydedilmiş."})
        connection.execute("INSERT INTO receipts (product_name,tax_id,invoice_no,receipt_no,receipt_datetime,vat_rate,vat_base,vat_amount,total_amount,created_at,user_id) VALUES (?,?,?,?,?,?,?,?,?,?,?)", (*[data[field] for field in RECEIPT_FIELDS[:5]], number(data["vat_rate"]), number(data["vat_base"]), number(data["vat_amount"]), number(data["total_amount"]), datetime.now().isoformat(timespec="seconds"), user_id))
    return jsonify({"ok": True})


@app.post("/api/ocr/receipt")
def receipt_ocr():
    text, error = ocr_text("image")
    result = extract_receipt(text) if text else {field: MISSING for field in RECEIPT_FIELDS}
    return jsonify({"data": result, "error": error, "engine": "PaddleOCR/EasyOCR"})


@app.get("/api/z-reports")
def z_reports():
    user_id = current_user_id()
    with get_db() as connection:
        rows = connection.execute("SELECT * FROM z_reports WHERE user_id=? ORDER BY id DESC", (user_id,)).fetchall()
    return jsonify([row_dict(row) for row in rows])


@app.post("/api/z-reports")
def create_z_report():
    user_id = current_user_id()
    data = payload(Z_INPUT_FIELDS)
    if not data["report_no"] or data["daily_turnover"] == "":
        return jsonify({"error": "Z raporu numarası ve günlük toplam ciro zorunludur."}), 400
    numeric = {field: number(data[field]) for field in Z_FIELDS if field not in {"report_no", "report_datetime", "first_receipt_no", "last_receipt_no"}}
    values = [data["report_no"], data["report_datetime"], numeric["daily_turnover"]]
    values += [numeric[f"vat{rate}_{kind}"] for rate in (1, 10, 20) for kind in ("base", "amount")]
    values += [numeric[field] for field in ("cash_amount", "card_amount", "other_payment", "cancel_amount", "refund_amount", "discount_amount")]
    values += [data["first_receipt_no"], data["last_receipt_no"], datetime.now().isoformat(timespec="seconds")]
    with get_db() as connection:
        values = values[:-1] + [data["product_name"], data["tax_id"], data["invoice_no"], values[-1]]
        connection.execute("INSERT INTO z_reports (report_no,report_datetime,daily_turnover,vat1_base,vat1_amount,vat10_base,vat10_amount,vat20_base,vat20_amount,cash_amount,card_amount,other_payment,cancel_amount,refund_amount,discount_amount,first_receipt_no,last_receipt_no,product_name,tax_id,invoice_no,created_at,user_id) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", values + [user_id])
    return jsonify({"ok": True})


@app.post("/api/ocr/z-report")
def z_report_ocr():
    text, error = ocr_text("image")
    result = extract_z_report(text) if text else {field: MISSING for field in Z_FIELDS}
    return jsonify({"data": result, "error": error, "engine": "PaddleOCR/EasyOCR"})


@app.get("/export/excel")
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    export_type = request.args.get("type", "all")
    user_id = current_user_id()
    if export_type not in {"all", "receipts", "z-reports"}:
        return jsonify({"error": "Geçersiz Excel çıktı türü."}), 400
    with get_db() as connection:
        receipt_rows = connection.execute("SELECT * FROM receipts WHERE user_id=? ORDER BY receipt_datetime", (user_id,)).fetchall()
        z_rows = connection.execute("SELECT * FROM z_reports WHERE user_id=? ORDER BY report_datetime", (user_id,)).fetchall()
    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="173F5F")

    def sheet(name: str, headers: list[str], rows: list[sqlite3.Row], keys: list[str] | None = None) -> None:
        keys = keys or headers
        ws = workbook.create_sheet(name)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        for row in rows:
            values = []
            for key in keys:
                if key == "payment_method":
                    methods = []
                    if row["cash_amount"]:
                        methods.append("Nakit")
                    if row["card_amount"]:
                        methods.append("Kredi Kartı / POS")
                    if row["other_payment"]:
                        methods.append("Diğer")
                    values.append(", ".join(methods) or MISSING)
                elif key == "vat_rate":
                    if "vat_rate" in row.keys():
                        values.append(row["vat_rate"] or MISSING)
                    else:
                        rates = [str(rate) for rate in (1, 10, 20) if row[f"vat{rate}_amount"] or row[f"vat{rate}_base"]]
                        values.append(", ".join(f"%{rate}" for rate in rates) or MISSING)
                elif key == "vat_amount":
                    if "vat_amount" in row.keys():
                        values.append(row["vat_amount"] or 0)
                    else:
                        values.append(sum((row[f"vat{rate}_amount"] or 0) for rate in (1, 10, 20)))
                else:
                    values.append(row[key])
            ws.append(values)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for column in ws.columns:
            ws.column_dimensions[column[0].column_letter].width = min(max(len(str(cell.value or "")) for cell in column) + 2, 28)

    if export_type in {"all", "receipts"}:
        sheet("Normal Fişler", ["Tarih-Saat", "Ürün/Hizmet Adı", "VKN/TCKN", "Belge No", "KDV Oranı", "KDV Tutarı", "Toplam Tutar"], receipt_rows, ["receipt_datetime", "product_name", "tax_id", "invoice_no", "vat_rate", "vat_amount", "total_amount"])
    if export_type in {"all", "z-reports"}:
        sheet("Z Raporları", ["Tarih-Saat", "Ürün/Hizmet Adı", "VKN/TCKN", "Belge No", "KDV Oranı", "KDV Tutarı", "Ödeme Yöntemi", "Toplam Tutar"], z_rows, ["report_datetime", "product_name", "tax_id", "report_no", "vat_rate", "vat_amount", "payment_method", "daily_turnover"])
    if export_type != "all":
        filename = "normal_fisler.xlsx" if export_type == "receipts" else "z_raporlari.xlsx"
        output = BASE_DIR / "data" / filename
        workbook.save(output)
        return send_file(output, as_attachment=True, download_name=filename)
    summary = workbook.create_sheet("Aylık Özet")
    summary.append(["Ay", "Fiş Cirosu", "Z Cirosu", "Toplam Ciro"])
    with get_db() as connection:
        months = connection.execute("SELECT month, SUM(receipt_total) AS receipt_total, SUM(z_total) AS z_total FROM (SELECT strftime('%Y-%m', receipt_datetime) AS month, total_amount AS receipt_total, 0 AS z_total FROM receipts WHERE receipt_datetime != '' AND user_id=? UNION ALL SELECT strftime('%Y-%m', report_datetime) AS month, 0 AS receipt_total, daily_turnover AS z_total FROM z_reports WHERE report_datetime != '' AND user_id=?) WHERE month IS NOT NULL GROUP BY month ORDER BY month", (user_id, user_id)).fetchall()
    for row in months:
        summary.append([row["month"], row["receipt_total"], row["z_total"], row["receipt_total"] + row["z_total"]])
    sheet("KDV Özeti", ["KDV Oranı", "Matrah", "KDV Tutarı"], [])
    vat = workbook["KDV Özeti"]
    for rate in (1, 10, 20):
        z_base = sum((row[f"vat{rate}_base"] or 0) for row in z_rows)
        z_amount = sum((row[f"vat{rate}_amount"] or 0) for row in z_rows)
        receipt_base = sum((row["vat_base"] or 0) for row in receipt_rows if number(row["vat_rate"]) == rate)
        receipt_amount = sum((row["vat_amount"] or 0) for row in receipt_rows if number(row["vat_rate"]) == rate)
        vat.append([f"%{rate}", z_base + receipt_base, z_amount + receipt_amount])
    sheet("Ödeme Özeti", ["Ödeme Tipi", "Tutar"], [])
    payment = workbook["Ödeme Özeti"]
    payment.append(["Nakit", sum((row["cash_amount"] or 0) for row in z_rows)])
    payment.append(["Kredi Kartı / POS", sum((row["card_amount"] or 0) for row in z_rows)])
    payment.append(["Diğer", sum((row["other_payment"] or 0) for row in z_rows)])
    output = BASE_DIR / "data" / f"fis_raporu_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    workbook.save(output)
    return send_file(output, as_attachment=True, download_name=output.name)


init_db()

if __name__ == "__main__":
    app.run(debug=True, port=int(os.getenv("PORT", "5000")))